import os
import openpyxl
from xlrd import open_workbook
global book
book=openpyxl.Workbook()
root='C:\Python27\Libtest3'
#ws1=book.create_sheet('main')
import random
book.save('total0715.xlsx')
for root, dirs,files in os.walk(root):
    
    book1=openpyxl.Workbook('total0715.xlsx')
    shs=book1.sheetnames
    for d in dirs:
        #os.chdir(os.path.join(root,d))
    
        print(os.path.join(root,d))
        #print 2
        di=os.path.join(root,d)
        print di
        l=os.listdir(di)
        shs=[str(i) for i in shs]
        r1=os.path.split(di)
        print r1
        r2=os.path.split(r1[-2])
        print r1[-1]
        print r2[-1]
        if d in shs:
            #sheet1=book.create_sheet(d+str(random.randint(0,9)))
            sheet1=book.create_sheet(r2[-1]+r1[-1])
            print 2
            for r in range(0,len(l)):
                sheetl.cell(row=r+1,column=1).value=l[r]
        else:
            sheetl = book.create_sheet(d)
            for r in range(0,len(l)):
                sheetl.cell(row=r+1,column=1).value=l[r]
        book.save('total0715.xlsx')
    book.save('total0715.xlsx')
    print book.sheetnames




##>>> os.path.split(root)
##('C:\\Python', 'Lib')
##>>> r1=os.path.split(root)
##>>> r2=os.path.split(r1[-2])
##>>> r2
##('C:\\', 'Python')
##>>> r1+r2
##('C:\\Python', 'Lib', 'C:\\', 'Python')
##>>> r1[-1]+r2[-1]
##'LibPython

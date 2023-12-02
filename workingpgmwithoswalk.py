import os
import sys
import openpyxl
import pandas as pd
import xlwt
from xlrd import open_workbook
from xlutils.copy import copy
import random
import re

book=openpyxl.Workbook()

#book.save('total24.xlsx')
root='C:\Python27\phoebebackup'
f=os.listdir('C:\Python27\phoebebackup')
print f
ws1=book.create_sheet('main')
for r in range(0,len(f)):
    ws1.cell(row=r+1,column=1).value=f[r]
    
book.save('total071519.xlsx')
for root, dirs,files in os.walk(root):

    for d in dirs:
        shs = book.sheetnames
        #print shs
        #print d
        l=os.listdir(os.path.join(root,d))
        #print(os.path.join(root,d))
        #print 2
        di=os.path.join(root,d)
        #print di
        l=os.listdir(di)
        shs=[str(i) for i in shs]
        r1=os.path.split(di)
        #print r1
        r2=os.path.split(r1[-2])
        #print r1[-1]
        #print r2[-1]
        ex=('.pyc')
        #l=[p for p in l if (p.endswith(ex) or os.path.isdir(p))]
        l=[p for p in l if not(p.endswith(ex))]
        shs=[str(i) for i in shs]
        #print shs
        if d in shs:
            sheet1=book.create_sheet(r1[-1]+r2[-1])
            print d
            print l
        else:
            sheetl = book.create_sheet(d)
        for r in range(0,len(l)):
            sheetl.cell(row=r+1,column=1).value=l[r]
##            ex=('.py')
##            if l[r].endswith(ex):
##                sheetl.cell(row=r+1,column=1).value=l[r].strip()
##                
    book.save('total071519.xlsx')
    exit


